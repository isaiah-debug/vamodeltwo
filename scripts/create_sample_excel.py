"""
Creates a sample Excel file you can use to test the pipeline immediately.
Run once:  python scripts/create_sample_excel.py
Output:    data/sample_dialogue.xlsx
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SAMPLE_TURNS = [
    ("Alice",  "Okay I think we should start by dividing the tasks clearly.", "Directive"),
    ("Bob",    "Yeah that makes sense. Maybe Alice can lead the analysis section?", "Agreement"),
    ("Carol",  "I don't know, I think we need more time to figure out the structure first.", "Hedging"),
    ("Alice",  "No, we have a deadline. Let's just move forward.", "Assertive"),
    ("David",  "Building on Alice's point — I can take the data collection part.", "Agreement"),
    ("Bob",    "Actually wait, I was going to say we should vote on the structure.", "Interrupt"),
    ("Alice",  "I really think voting will just slow us down.", "Directive"),
    ("Carol",  "Maybe Alice is right. I'm okay with her leading this.", "Agreement"),
    ("David",  "Sounds good. So Alice leads, I do data, Bob does what?", "Directive"),
    ("Bob",    "I guess I can do the write-up.", "Hedging"),
    ("Alice",  "Perfect. Carol, can you handle the literature review?", "Directive"),
    ("Carol",  "Sure, I can do that.", "Agreement"),
    ("Alice",  "Great, we'll check in on Thursday then.", "Directive"),
    ("David",  "Thursday works for me.", "Agreement"),
    ("Bob",    "Hold on — what if we fall behind? Should we have a backup plan?", "Concern"),
    ("Alice",  "We won't fall behind if everyone does their part.", "Assertive"),
    ("Carol",  "I think Bob has a point. A backup plan isn't a bad idea.", "Agreement"),
    ("David",  "I agree with Carol. Let's say if we're not done by Wednesday we regroup.", "Directive"),
    ("Alice",  "Fine. Wednesday checkpoint. Everyone okay with that?", "Agreement"),
    ("Bob",    "Yes, works for me.", "Agreement"),
    ("Carol",  "Me too.", "Agreement"),
    ("David",  "Perfect. I think we're all set.", "Assertive"),
]

def create():
    Path("data").mkdir(exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Session 01"

    # Header
    headers = ["turn", "speaker", "utterance", "code"]
    header_fill = PatternFill("solid", fgColor="2D2D5E")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    for i, (spk, utt, code) in enumerate(SAMPLE_TURNS, 1):
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=spk)
        ws.cell(row=i+1, column=3, value=utt)
        ws.cell(row=i+1, column=4, value=code)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 65
    ws.column_dimensions["D"].width = 15

    out = Path("data/sample_dialogue.xlsx")
    wb.save(out)
    print(f"Sample Excel created: {out}")
    print(f"  {len(SAMPLE_TURNS)} turns, 4 speakers (Alice, Bob, Carol, David)")
    print(f"  Upload this file in the dashboard to test the pipeline.")

if __name__ == "__main__":
    create()
