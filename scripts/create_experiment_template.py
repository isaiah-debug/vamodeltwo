"""
Creates the dialogue coding Excel template for the escape room experiment.
Players: A, B, C, D  |  ~29 min session  |  Headcam + 3 security cameras

Run: python scripts/create_experiment_template.py
Output: data/escape_room_dialogue_template.xlsx  (blank, ready to fill in)
        data/escape_room_dialogue_EXAMPLE.xlsx   (5 prefilled rows to show format)
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

PLAYERS = ["A", "B", "C", "D"]

# Behavior codes from the codebook — edit to match yours
BEHAVIOR_CODES = [
    "Directive",
    "Agreement",
    "Disagreement",
    "Suggestion",
    "Question",
    "Information",
    "Hedging",
    "Interruption",
    "Acknowledgment",
    "Problem-solving",
    "Humor",
    "Silence/Pause",
    "Other",
]

PLAYER_COLORS = {
    "A": "C6EFCE",   # green
    "B": "BDD7EE",   # blue
    "C": "FCE4D6",   # peach
    "D": "EAD1DC",   # pink
}

def col_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def header_style(ws, row, cols, text_list, bg="1F3864", fg="FFFFFF"):
    hf = Font(bold=True, color=fg)
    hfill = PatternFill("solid", fgColor=bg)
    for col, text in zip(cols, text_list):
        c = ws.cell(row=row, column=col, value=text)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def build_template(wb, sheet_name, prefill_rows=None):
    ws = wb.create_sheet(sheet_name)

    # ── Session header block ────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    title = ws["A1"]
    title.value = "ESCAPE ROOM DIALOGUE CODING SHEET"
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="1F3864")
    title.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    meta_fields = [
        ("A2", "Session ID:", "B2", ""),
        ("C2", "Date:", "D2", ""),
        ("E2", "Coder:", "F2", ""),
        ("G2", "Camera source:", "H2", "Security cam 1 / 2 / 3 / Headcam A / B / C / D"),
    ]
    for label_cell, label_val, val_cell, hint in meta_fields:
        ws[label_cell].value = label_val
        ws[label_cell].font = Font(bold=True)
        ws[val_cell].value = hint
        ws[val_cell].font = Font(italic=True, color="808080")

    ws.merge_cells("A3:J3")
    sync = ws["A3"]
    sync.value = (
        "SYNC NOTE: t=0 is Player A clap + Owen says 'Have at it!'  "
        "Fill start_offset (seconds) per player before coding."
    )
    sync.font = Font(italic=True, color="7030A0")
    sync.fill = PatternFill("solid", fgColor="F2E7FE")
    ws.row_dimensions[3].height = 18

    # Player start offset row
    ws["A4"].value = "Player start offset (s):"
    ws["A4"].font = Font(bold=True)
    for i, p in enumerate(PLAYERS, 1):
        ws.cell(row=4, column=i+1).value = f"Player {p}:"
        ws.cell(row=4, column=i+1).font = Font(bold=True, color="404040")
        ws.cell(row=4, column=i+2-1).fill = col_fill(PLAYER_COLORS[p])

    ws.row_dimensions[5].height = 6   # spacer

    # ── Column headers ──────────────────────────────────────────────────
    COLS = [
        "turn",          # A=1
        "timestamp_s",   # B=2  seconds from t=0
        "speaker",       # C=3
        "utterance",     # D=4
        "behavior_code", # E=5
        "directed_to",   # F=6  who are they addressing?
        "overlap",       # G=7  Y/N overlap/interrupt
        "notes",         # H=8
        "cam_source",    # I=9
        "confidence",    # J=10  coder confidence 1-3
    ]
    header_style(ws, 6, range(1, 11), COLS)
    ws.row_dimensions[6].height = 22

    # Sub-header hints
    hints = [
        "#", "sec from t=0", "A/B/C/D", "exact words spoken",
        "/".join(BEHAVIOR_CODES[:4]) + "…",
        "A/B/C/D/all", "Y/N",
        "observations", "sec cam 1/2/3 or headcam X",
        "1=certain 2=likely 3=guess"
    ]
    for col, hint in enumerate(hints, 1):
        c = ws.cell(row=7, column=col, value=hint)
        c.font = Font(italic=True, size=8, color="606060")
        c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[7].height = 30

    # ── Data rows ───────────────────────────────────────────────────────
    START_ROW = 8
    NUM_BLANK = 200

    example_data = prefill_rows or []

    for i in range(NUM_BLANK):
        row = START_ROW + i
        if i < len(example_data):
            values = example_data[i]
        else:
            values = [i+1] + [""] * 9

        player = str(values[2]) if len(values) > 2 else ""
        fill = col_fill(PLAYER_COLORS.get(player, "FFFFFF"))

        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", wrap_text=(col == 4))
            if player in PLAYER_COLORS:
                c.fill = col_fill(PLAYER_COLORS[player] + "88"[:0])   # subtle
        ws.row_dimensions[row].height = 16

    # ── Column widths ───────────────────────────────────────────────────
    widths = [6, 12, 9, 55, 18, 12, 8, 30, 20, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Code reference sheet ────────────────────────────────────────────
    return ws


def build_codebook(wb):
    ws = wb.create_sheet("Behavior Codes")
    ws["A1"].value = "Behavior Code Reference"
    ws["A1"].font = Font(bold=True, size=13)

    defs = {
        "Directive":      "Instructs or tells another player what to do",
        "Agreement":      "Explicitly agrees with or validates another's idea",
        "Disagreement":   "Explicitly disagrees or pushes back",
        "Suggestion":     "Proposes an idea without commanding ('maybe we could...')",
        "Question":       "Asks for information or clarification",
        "Information":    "Shares a fact, observation, or update",
        "Hedging":        "Qualifies or softens a statement ('I think', 'maybe')",
        "Interruption":   "Speaks over or cuts off another player",
        "Acknowledgment": "Brief feedback token: 'yeah', 'uh huh', 'okay'",
        "Problem-solving":"Directly works through a puzzle or obstacle",
        "Humor":          "Joke, laughter, or playful aside",
        "Silence/Pause":  "Coder-notable silence or hesitation (>2s)",
        "Other":          "Does not fit above; describe in notes",
    }

    ws.cell(row=2, column=1, value="Code").font = Font(bold=True)
    ws.cell(row=2, column=2, value="Definition").font = Font(bold=True)

    fills = [PatternFill("solid", fgColor=c) for c in
             ["E8F5E9","E3F2FD","FFF3E0","F3E5F5","FCE4D6","E0F7FA",
              "FFFDE7","F9FBE7","E8EAF6","FBE9E7","F1F8E9","E0F2F1","FAFAFA"]]

    for i, (code, defn) in enumerate(defs.items(), 3):
        ws.cell(row=i, column=1, value=code).fill = fills[i-3]
        ws.cell(row=i, column=2, value=defn)
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 60

    # Player color key
    ws.cell(row=18, column=1, value="Player Color Key").font = Font(bold=True)
    for i, (p, color) in enumerate(PLAYER_COLORS.items(), 19):
        c = ws.cell(row=i, column=1, value=f"Player {p}")
        c.fill = col_fill(color)
        c.font = Font(bold=True)


EXAMPLE_ROWS = [
    [1,  "0",    "A", "Okay let's go, look around the room.", "Directive",  "all", "N", "", "headcam_A", 1],
    [2,  "4",    "B", "There's numbers on this lock over here.", "Information","all","N","north wall padlock","sec_cam_1",1],
    [3,  "9",    "C", "Should we split up?", "Question", "all", "N", "", "headcam_C", 1],
    [4,  "11",   "A", "Yeah split up, cover more ground.", "Agreement", "C", "N", "", "headcam_A", 1],
    [5,  "13",   "D", "I'll take the desk.", "Directive", "all", "N", "", "sec_cam_2", 1],
    [6,  "15",   "B", "Wait I think these numbers match — 4, 7, 2?", "Problem-solving","A","N","trying lock","headcam_B",1],
    [7,  "20",   "A", "No no hold on, there's a clue on the wall.", "Interruption","B","Y","overlap w B","headcam_A",1],
    [8,  "24",   "C", "What does it say?", "Question", "A", "N", "", "headcam_C", 1],
    [9,  "26",   "A", "It says 'the order matters'. Maybe it's not 4-7-2.", "Information","all","N","","headcam_A",1],
    [10, "31",   "D", "Found a key in the drawer.", "Information", "all", "N", "small brass key", "sec_cam_2", 1],
]


if __name__ == "__main__":
    Path("data").mkdir(exist_ok=True)

    # Blank template
    wb_blank = openpyxl.Workbook()
    wb_blank.remove(wb_blank.active)
    build_template(wb_blank, "Dialogue Coding", prefill_rows=[[i+1]+[""]*9 for i in range(200)])
    build_codebook(wb_blank)
    blank_path = Path("data/escape_room_dialogue_TEMPLATE.xlsx")
    wb_blank.save(blank_path)
    print(f"Blank template -> {blank_path}")

    # Example with prefilled rows
    wb_ex = openpyxl.Workbook()
    wb_ex.remove(wb_ex.active)
    build_template(wb_ex, "Dialogue Coding", prefill_rows=EXAMPLE_ROWS)
    build_codebook(wb_ex)
    ex_path = Path("data/escape_room_dialogue_EXAMPLE.xlsx")
    wb_ex.save(ex_path)
    print(f"Example (10 prefilled rows) -> {ex_path}")
    print()
    print("COLUMNS:")
    print("  turn         — sequential number")
    print("  timestamp_s  — seconds from t=0 (Player A clap)")
    print("  speaker      — A / B / C / D")
    print("  utterance    — exact words")
    print("  behavior_code— see 'Behavior Codes' sheet")
    print("  directed_to  — A / B / C / D / all")
    print("  overlap      — Y if they talk over someone")
    print("  notes        — your observations")
    print("  cam_source   — which camera you pulled this from")
    print("  confidence   — 1=certain 2=likely 3=guess")
